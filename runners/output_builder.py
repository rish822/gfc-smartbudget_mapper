"""
Output workbook builder — shared across all project runners.
Produces the standard 8-sheet enrichment workbook.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
from engine.gfc_mapping_engine import classify_sheet

# ── Styles ────────────────────────────────────────────────────────────────────
BLUE   = "1F4E78"
LIGHT  = "F2F2F2"
GREEN  = "C6EFCE"
YELLOW = "FFEB9C"
RED    = "FFC7CE"
THIN   = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor=BLUE)
R_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _header(ws, row, n):
    for c in range(1, n+1):
        cell = ws.cell(row, c)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = CENTER; cell.border = BORDER


def _zebra(ws, r0, r1, n):
    for r in range(r0, r1+1):
        fill = PatternFill("solid", fgColor=LIGHT) if (r-r0) % 2 else None
        for c in range(1, n+1):
            cell = ws.cell(r, c)
            cell.font = R_FONT; cell.alignment = LEFT; cell.border = BORDER
            if fill: cell.fill = fill


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_output_workbook(results, budget, master, gfc_path, project_name, out_path) -> dict:
    """Build the 8-sheet enrichment workbook. Returns stats dict."""
    from engine.gfc_mapping_engine import MATCH_THRESHOLDS

    match_counts  = Counter(r.match_status for r in results)
    status_counts = Counter(r.norm_status  for r in results)
    pr_ready      = sum(1 for r in results if r.pr_ready)
    by_sheet      = defaultdict(list)
    for r in results: by_sheet[r.gfc_sheet].append(r)

    src_wb = openpyxl.load_workbook(gfc_path, read_only=True)
    sheet_audit = []
    for s in src_wb.sheetnames:
        cat, sub, conf, reason = classify_sheet(s)
        decision = "SKIP" if (cat is None and "SKIP" in reason) else ("MAPPED" if cat else "UNCLASSIFIED")
        sheet_audit.append({"Sheet": s, "Category": cat or "—", "Sub-cat": sub or "—",
                             "Decision": decision, "Items": len(by_sheet.get(s, [])), "Reason": reason})

    matched_boqs = {r.matched_boq_no for r in results if r.matched_boq_no}
    unlinked     = [b for b in budget if b["BOQ No"] not in matched_boqs]

    boq_cats = {(b.get("Category") or "").upper() for b in budget}
    new_scope = defaultdict(lambda: {"sheets": set(), "items": [], "count": 0})
    for r in results:
        if r.detected_category and r.detected_category.upper() not in boq_cats:
            ns = new_scope[r.detected_category]
            ns["sheets"].add(r.gfc_sheet); ns["count"] += 1
            if len(ns["items"]) < 5:
                ns["items"].append(r.raw_description or r.raw_areas or "(unnamed)")

    auto = match_counts.get("🟢 Matched", 0)
    sug  = match_counts.get("🟡 Suggested", 0)
    nib  = match_counts.get("🔴 Not in Budget", 0)
    tot  = max(auto + sug + nib, 1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("1. Run Summary")
    ws["A1"] = f"{project_name}  —  GFC ↔ Smart Budget Enrichment"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=BLUE)
    ws.merge_cells("A1:H1")

    ws["A3"] = "Headline Metrics"
    ws["A3"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
    metrics = [
        ("GFC sheets in file",                          len(src_wb.sheetnames)),
        ("  Sheets auto-classified to a Category",      sum(1 for s in sheet_audit if s["Decision"]=="MAPPED")),
        ("  Sheets skipped (Drawing Schedule / Index)", sum(1 for s in sheet_audit if s["Decision"]=="SKIP")),
        ("  Sheets unclassified (need rule)",           sum(1 for s in sheet_audit if s["Decision"]=="UNCLASSIFIED")),
        ("", ""),
        ("BOQ line items",                              len(budget)),
        ("GFC line items extracted",                    len(results)),
        ("", ""),
        (f"🟢 Auto-Matched  (score ≥ {MATCH_THRESHOLDS['auto']})",    auto),
        (f"🟡 Suggested     (score {MATCH_THRESHOLDS['suggested']}–{MATCH_THRESHOLDS['auto']-1})", sug),
        (f"🔴 Not in BOQ   (score < {MATCH_THRESHOLDS['suggested']})", nib),
        ("", ""),
        ("Status: Client Approved",                     status_counts.get("Client Approved", 0)),
        ("Status: Pending",                             status_counts.get("Pending", 0)),
        ("Status: Excluded",                            status_counts.get("Excluded", 0)),
        ("", ""),
        ("✅ PR-ready (matched + approved + qty + uom)", pr_ready),
        ("BOQ items WITH a GFC link",                   len(matched_boqs)),
        ("BOQ items WITHOUT a GFC link",                len(unlinked)),
    ]
    for i, (k, v) in enumerate(metrics, start=4):
        ws.cell(i, 1, k).font = Font(name="Arial", size=10, bold=bool(k and not k.startswith(" ")))
        ws.cell(i, 2, v).font = Font(name="Arial", size=10)
        if isinstance(v, int):
            ws.cell(i, 2).alignment = Alignment(horizontal="right")

    ws["A25"] = f"Auto-match: {100*auto/tot:.1f}%   |   With suggested: {100*(auto+sug)/tot:.1f}%"
    ws["A25"].font = Font(name="Arial", size=11, bold=True, color="2E7D32")

    interp = [
        ("🟢 Auto-Matched",  "High confidence — GFC row auto-linked to BOQ item. Spot-check exceptions only."),
        ("🟡 Suggested",     "Likely match. Human confirms or rejects in GFC Connect review screen (one click)."),
        ("🔴 Not in BOQ",    "(a) Genuine new scope → flag PM, price as Variation. OR (b) abstract label that survived row classification."),
        ("PR-ready",         "GF5.1 criteria met: matched + Client Approved + has qty + has UoM. Drop into PR."),
        ("BOQ Coverage Gap", "BOQ item with no GFC row. Civil/MEP/Services = expected. Furniture/Decor here = possible scope drop."),
    ]
    ws["A27"] = "Legend"
    ws["A27"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
    for i, (lbl, txt) in enumerate(interp, start=28):
        ws.cell(i, 1, lbl).font = Font(name="Arial", size=10, bold=True)
        ws.cell(i, 2, txt).font = Font(name="Arial", size=10)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        ws.row_dimensions[i].height = 28
    _widths(ws, [55, 12])

    # ── Sheet 2: Enriched GFC Items ───────────────────────────────────────────
    ws = wb.create_sheet("2. Enriched GFC Items")
    hdrs = ["GFC Sheet","Row","Floor","Row Type",
            "GFC Description","GFC Brand","GFC Areas","GFC Size",
            "Design Qty","Ops Qty","Norm UoM","Raw Status","Norm Status",
            "Master Category","Master Sub-cat","Master Item",
            "Match Status","Score",
            "BOQ No","BOQ Description","BOQ Brand","BOQ UoM","BOQ Unit Rate",
            "PR Ready","PR Blocker"]
    ws.append(hdrs); _header(ws, 1, len(hdrs))
    for r in results:
        ws.append([r.gfc_sheet, r.gfc_row_idx, r.floor_label, r.row_type,
                   r.raw_description, r.raw_brand, r.raw_areas, r.raw_size,
                   r.raw_qty_design, r.raw_qty_ops, r.norm_uom,
                   r.raw_status, r.norm_status,
                   r.master_category, r.master_subcategory, r.master_item,
                   r.match_status, r.match_score,
                   r.matched_boq_no, r.matched_description, r.matched_brand,
                   r.matched_uom, r.matched_bcs_rate,
                   "✅" if r.pr_ready else "❌", r.pr_blocker])
    _zebra(ws, 2, ws.max_row, len(hdrs))
    for ri in range(2, ws.max_row+1):
        sc = ws.cell(ri, 17)
        if sc.value == "🟢 Matched":      sc.fill = PatternFill("solid", fgColor=GREEN)
        elif sc.value == "🟡 Suggested":  sc.fill = PatternFill("solid", fgColor=YELLOW)
        elif sc.value == "🔴 Not in Budget": sc.fill = PatternFill("solid", fgColor=RED)
        pr = ws.cell(ri, 24)
        if pr.value == "✅": pr.fill = PatternFill("solid", fgColor=GREEN)
    ws.freeze_panes = "A2"
    _widths(ws, [22,7,7,9,42,16,20,16,9,9,9,16,14,18,20,22,16,7,10,38,16,9,12,9,30])

    # ── Sheet 3: Score Breakdown ───────────────────────────────────────────────
    ws = wb.create_sheet("3. Score Breakdown")
    hdrs = ["GFC Sheet","Row","GFC Description","Match Status","Score","BOQ No","Breakdown"]
    ws.append(hdrs); _header(ws, 1, len(hdrs))
    for r in results:
        ws.append([r.gfc_sheet, r.gfc_row_idx, r.raw_description,
                   r.match_status, r.match_score, r.matched_boq_no or "—", r.score_breakdown])
    _zebra(ws, 2, ws.max_row, len(hdrs)); ws.freeze_panes = "A2"
    _widths(ws, [22, 7, 42, 16, 7, 12, 110])

    # ── Sheet 4: L1 Sheet Audit ────────────────────────────────────────────────
    ws = wb.create_sheet("4. L1 Sheet Audit")
    hdrs = ["Sheet Name","Detected Category","Sub-cat Hint","Decision","Items Extracted","Reason"]
    ws.append(hdrs); _header(ws, 1, len(hdrs))
    for s in sheet_audit:
        ws.append([s["Sheet"], s["Category"], s["Sub-cat"], s["Decision"], s["Items"], s["Reason"]])
    _zebra(ws, 2, ws.max_row, len(hdrs))
    for ri in range(2, ws.max_row+1):
        d = ws.cell(ri, 4).value
        fill = {"MAPPED": GREEN, "SKIP": LIGHT, "UNCLASSIFIED": YELLOW}.get(d)
        if fill: ws.cell(ri, 4).fill = PatternFill("solid", fgColor=fill)
    _widths(ws, [32, 22, 22, 14, 12, 55])

    # ── Sheet 5: BOQ Coverage Gaps ────────────────────────────────────────────
    ws = wb.create_sheet("5. BOQ Coverage Gaps")
    ws["A1"] = "BOQ items with NO matching GFC line item"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
    ws["A2"] = "Civil/MEP/Services = expected. Any Furniture/Decor/Flooring here = possible scope drop — alert design team."
    ws["A2"].font = Font(name="Arial", size=10, italic=True); ws.merge_cells("A2:I2")
    hdrs = ["BOQ No","Section","Category","Sub-category","Description","Make/Brand","UoM","BCS Qty","Likely Reason"]
    for c, h in enumerate(hdrs, 1): ws.cell(4, c, h)
    _header(ws, 4, len(hdrs))
    for b in unlinked:
        reason = ("Civil/MEP/Services — no GFC expected" if b.get("Category") == "CIVIL"
                  else "⚠️ No GFC match — check with design team")
        ws.append([b["BOQ No"], b.get("Section",""), b.get("Category",""),
                   b.get("Sub-category",""), b["Description"], b.get("Make/Brand",""),
                   b.get("UoM",""), b.get("BCS Qty",""), reason])
    _zebra(ws, 5, ws.max_row, len(hdrs))
    _widths(ws, [10, 30, 18, 22, 42, 22, 8, 10, 38])

    # ── Sheet 6: New Scope (GF6.5) ────────────────────────────────────────────
    ws = wb.create_sheet("6. New Scope (GF6.5)")
    ws["A1"] = "GFC categories ABSENT from BOQ — potential new scope"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=BLUE)
    ws.merge_cells("A1:G1")
    hdrs = ["GFC Category","GFC Sheets","Item Count","Sample GFC Items","Recommendation"]
    for c, h in enumerate(hdrs, 1): ws.cell(3, c, h)
    _header(ws, 3, len(hdrs))
    for cat, info in sorted(new_scope.items(), key=lambda x: -x[1]["count"]):
        ws.append([cat, ", ".join(sorted(info["sheets"])), info["count"],
                   " · ".join(info["items"][:4]),
                   "⚠️ Flag with PM — price as Variation or correct BOQ"])
    if not new_scope:
        ws.append(["(none detected)", "—", 0, "—", "✅ All GFC categories have BOQ counterparts"])
    _zebra(ws, 4, ws.max_row, len(hdrs))
    _widths(ws, [18, 32, 10, 55, 42])

    # ── Sheet 7: BOQ Parsed ───────────────────────────────────────────────────
    ws = wb.create_sheet("7. BOQ Parsed")
    hdrs = ["BOQ No","Section","Category","Sub-category","Description",
            "Specification","Make/Brand","UoM","BCS Qty","BCS Rate","BCS Amount","Space/Area"]
    ws.append(hdrs); _header(ws, 1, len(hdrs))
    for b in budget:
        ws.append([b["BOQ No"], b.get("Section",""), b.get("Category",""),
                   b.get("Sub-category",""), b["Description"],
                   b.get("Specification",""), b.get("Make/Brand",""),
                   b.get("UoM",""), b.get("BCS Qty",""), b.get("BCS Rate",""),
                   b.get("BCS Amount",""), b.get("Space/Area","")])
    _zebra(ws, 2, ws.max_row, len(hdrs)); ws.freeze_panes = "A2"
    _widths(ws, [10, 30, 18, 22, 42, 38, 22, 8, 10, 10, 12, 25])

    # ── Sheet 8: Category Master ──────────────────────────────────────────────
    ws = wb.create_sheet("8. Category Master")
    ws.append(["Category","Sub-category","Item","Item Type"]); _header(ws, 1, 4)
    for m in master:
        ws.append([m["Category"], m["Sub-category"], m["Item"], m["Item Type"]])
    _zebra(ws, 2, ws.max_row, 4); _widths(ws, [22, 28, 36, 12])

    wb.save(out_path)

    stats = {
        "total_sheets": len(src_wb.sheetnames),
        "mapped_sheets": sum(1 for s in sheet_audit if s["Decision"]=="MAPPED"),
        "skipped_sheets": sum(1 for s in sheet_audit if s["Decision"]=="SKIP"),
        "unclassified_sheets": sum(1 for s in sheet_audit if s["Decision"]=="UNCLASSIFIED"),
        "auto": auto, "suggested": sug, "not_in_budget": nib,
        "pr_ready": pr_ready,
        "boq_linked": len(matched_boqs),
        "new_scope": list(new_scope.keys()),
    }
    return stats
