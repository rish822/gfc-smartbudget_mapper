"""
Unified, auto-detecting BOQ loader.
Works with any BOQ Excel regardless of structure — no format selection needed.
  - Multi-sheet: one sheet per category (Furniture / Flooring / etc.)
  - Single-sheet: section codes in col A (A, B, C... or G1, G2...)
Column mapping is driven entirely by header keyword matching, not hard-coded positions.
"""
import openpyxl
import re

# ── Category resolution ───────────────────────────────────────────────────────

# Sheet name token → Category (multi-sheet BOQ mode)
# Aligned with Category Master — fabrication→CIVIL, skirting/transition→DECORATIVES
SHEET_CATEGORY_TOKENS = [
    # FURNITURE
    (("furniture",),               "FURNITURE"),
    (("loose",),                   "FURNITURE"),
    (("chair",),                   "FURNITURE"),
    (("carpentry",),               "FURNITURE"),
    (("carpentary",),              "FURNITURE"),
    (("caprentry",),               "FURNITURE"),
    (("carpentery",),              "FURNITURE"),
    (("modular",),                 "FURNITURE"),
    (("rack",),                    "FURNITURE"),
    # FLOORING
    (("flooring",),                "FLOORING"),
    (("tile",),                    "FLOORING"),
    (("carpet",),                  "FLOORING"),
    (("spc",),                     "FLOORING"),
    (("lvt",),                     "FLOORING"),
    (("antistatic",),              "FLOORING"),
    (("anti", "static"),           "FLOORING"),
    # CEILING
    (("ceiling",),                 "CEILING"),
    (("open", "cell"),             "CEILING"),
    # PAINT
    (("paint",),                   "PAINT"),
    # ACOUSTIC
    (("acoustic",),                "ACOUSTIC"),
    (("accoustic",),               "ACOUSTIC"),
    # LIGHTING
    (("light",),                   "LIGHTING"),
    (("electrical",),              "LIGHTING"),
    # PARTITIONS AND DOORS
    (("partition",),               "PARTITIONS AND DOORS"),
    (("door",),                    "PARTITIONS AND DOORS"),
    (("glazing",),                 "PARTITIONS AND DOORS"),
    (("phenolic",),                "PARTITIONS AND DOORS"),
    # DECORATIVES — skirting/transition moved here per Category Master
    (("decoratives",),             "DECORATIVES"),
    (("blind",),                   "DECORATIVES"),
    (("wallpaper",),               "DECORATIVES"),
    (("fluted",),                  "DECORATIVES"),
    (("logo",),                    "DECORATIVES"),
    (("signage",),                 "DECORATIVES"),
    (("branding",),                "DECORATIVES"),
    (("frame",),                   "DECORATIVES"),
    (("decal",),                   "DECORATIVES"),
    (("fabric",),                  "DECORATIVES"),
    (("canvas",),                  "DECORATIVES"),
    (("panelling",),               "DECORATIVES"),
    (("cladding",),                "DECORATIVES"),
    (("skirting",),                "DECORATIVES"),
    (("transition",),              "DECORATIVES"),
    # SURFACE AND FINISHES
    (("surface",),                 "SURFACE AND FINISHES"),
    (("finish",),                  "SURFACE AND FINISHES"),
    (("laminate",),                "SURFACE AND FINISHES"),
    (("veneer",),                  "SURFACE AND FINISHES"),
    # CIVIL — fabrication moved here per Category Master
    (("civil",),                   "CIVIL"),
    (("services",),                "CIVIL"),
    (("hvac",),                    "CIVIL"),
    (("plumbing",),                "CIVIL"),
    (("sanitary",),                "CIVIL"),
    (("dado",),                    "CIVIL"),
    (("counter",),                 "CIVIL"),
    (("fabricat",),                "CIVIL"),
    (("ms", "fab"),                "CIVIL"),
]

# Sheets to skip in multi-sheet mode
SKIP_SHEET_TOKENS = (
    "summary", "total", "index", "cover", "legend", "timeline",
    "mom", "notes", "intro", "instruction", "terms", "sheet1",
    "drawing schedule", "tender", "tv size", "cumulative", "3ds",
)

# Single-letter section code → Category (single-sheet mode)
SECTION_CODE_CATEGORY = {
    "A": "FURNITURE",
    "B": "FLOORING",
    "C": "PARTITIONS AND DOORS",
    "D": "CIVIL",
    "E": "CEILING",
    "F": "DECORATIVES",
    "G": "CIVIL",
    "H": "LIGHTING",
    "I": "CIVIL",
}

# Sub-section code → Sub-category (G1/G2/C1 style in multi-sheet BOQs)
# Values use exact Category Master sub-category names.
SUBSECTION_CODE_SUBCAT = {
    "G1": "MODULAR FURNITURE",
    "G2": "LOOSE FURNITURE",
    "G3": "CUSTOM FURNITURE",
    "C1": "GYPSUM CEILING",
    "C2": "METAL CEILING",        # master: METAL CEILING (was GRID CEILING)
    "C3": "DECORATIVE CEILING",
    "D1": "TILE FLOORING",        # master: TILE FLOORING (was generic FLOORING)
    "E1": "GLASS PARTITION",
    "B1": "GRAPHICS",             # master: GRAPHICS (was WALL)
    "B2": "METAL CEILING",        # master: METAL CEILING (was generic CEILING)
    "F1": "PANELLING",            # master: PANELLING (was generic DECORATIVES)
    "F2": "UTILITIES",
}

# Description keyword → Category override (both modes)
# Skirting/transition now live under DECORATIVES per Category Master.
# Fabrication now lives under CIVIL per Category Master.
KEYWORD_CATEGORY = [
    (("blind",),             "DECORATIVES"),
    (("skirting",),          "DECORATIVES"),        # master: DECORATIVES / SKIRTING AND PROFILE
    (("transition", "profile"), "DECORATIVES"),     # master: DECORATIVES / SKIRTING AND PROFILE
    (("branding",),          "DECORATIVES"),
    (("signage",),           "DECORATIVES"),
    (("wallpaper",),         "DECORATIVES"),
    (("glass", "film"),      "DECORATIVES"),
    (("fluted", "panel"),    "DECORATIVES"),
    (("wall", "cladding"),   "DECORATIVES"),
    (("demolition",),        "CIVIL"),
    (("debris",),            "CIVIL"),
    (("siporex",),           "CIVIL"),
    (("pcc",),               "CIVIL"),
    (("waterproof",),        "CIVIL"),
    (("dado", "tile"),       "CIVIL"),
    (("counter", "top"),     "CIVIL"),
    (("fabricat",),          "CIVIL"),              # master: CIVIL / FABRICATION
    (("raised", "floor"),    "FLOORING"),
    (("antistatic",),        "FLOORING"),
    (("vitrified",),         "FLOORING"),
    (("carpet",),            "FLOORING"),
    (("tile",),              "FLOORING"),
    (("paint",),             "PAINT"),
    (("punning",),           "PAINT"),
    (("ceiling",),           "CEILING"),
    (("led", "light"),       "LIGHTING"),
    (("light", "led"),       "LIGHTING"),
    (("laminate",),          "SURFACE AND FINISHES"),
    (("veneer",),            "SURFACE AND FINISHES"),
]

# Description keyword → Sub-category (both modes)
# All values use exact Category Master sub-category names.
KEYWORD_SUBCAT = [
    # ── FURNITURE ─────────────────────────────────────────────────────────
    (("workstation",),              "MODULAR FURNITURE"),
    (("workstation", "partition"),  "MODULAR FURNITURE"),
    (("meeting", "table"),          "MODULAR FURNITURE"),
    (("conference", "table"),       "MODULAR FURNITURE"),
    (("cafeteria", "table"),        "MODULAR FURNITURE"),
    (("cafe", "table"),             "MODULAR FURNITURE"),
    (("cafeteria",),                "MODULAR FURNITURE"),
    (("phone", "booth"),            "MODULAR FURNITURE"),
    (("height", "adjustable"),      "MODULAR FURNITURE"),
    (("folding", "table"),          "MODULAR FURNITURE"),
    (("storage", "modular"),        "MODULAR FURNITURE"),
    (("cabin", "table"),            "CUSTOM FURNITURE"),
    (("reception", "desk"),         "CUSTOM FURNITURE"),
    (("reception", "table"),        "CUSTOM FURNITURE"),
    (("host", "desk"),              "CUSTOM FURNITURE"),
    (("book", "shelf"),             "CUSTOM FURNITURE"),
    (("wardrobe",),                 "CUSTOM FURNITURE"),
    (("pantry",),                   "CUSTOM FURNITURE"),
    (("booth", "seating"),          "CUSTOM FURNITURE"),
    (("step", "seating"),           "CUSTOM FURNITURE"),
    (("ledge", "seating"),          "CUSTOM FURNITURE"),
    (("low height storage",),       "CUSTOM FURNITURE"),
    (("custom", "storage"),         "CUSTOM FURNITURE"),
    (("storage",),                  "CUSTOM FURNITURE"),
    (("sofa",),                     "LOOSE FURNITURE"),
    (("centre table",),             "LOOSE FURNITURE"),
    (("center table",),             "LOOSE FURNITURE"),
    (("coffee", "table"),           "LOOSE FURNITURE"),
    (("side table",),               "LOOSE FURNITURE"),
    (("pouffe",),                   "LOOSE FURNITURE"),
    (("ottoman",),                  "LOOSE FURNITURE"),
    (("ottomon",),                  "LOOSE FURNITURE"),
    (("bar", "stool"),              "LOOSE FURNITURE"),
    (("lounger",),                  "LOOSE FURNITURE"),
    (("lounge", "chair"),           "LOOSE FURNITURE"),   # master: LOOSE FURNITURE
    (("cafe", "chair"),             "LOOSE FURNITURE"),   # master: LOOSE FURNITURE
    (("dining", "chair"),           "LOOSE FURNITURE"),
    (("dining", "table"),           "LOOSE FURNITURE"),
    (("bench",),                    "LOOSE FURNITURE"),
    (("pod", "seating"),            "LOOSE FURNITURE"),
    (("high back chair",),          "CHAIRS"),
    (("medium back chair",),        "CHAIRS"),
    (("executive chair",),          "CHAIRS"),
    (("task chair",),               "CHAIRS"),
    (("office chair",),             "CHAIRS"),
    (("visitor chair",),            "CHAIRS"),
    (("training chair",),           "CHAIRS"),
    (("gaming chair",),             "CHAIRS"),
    (("chair",),                    "CHAIRS"),

    # ── FLOORING ──────────────────────────────────────────────────────────
    (("vitrified", "tile"),         "TILE FLOORING"),
    (("ceramic", "tile"),           "TILE FLOORING"),
    (("porcelain", "tile"),         "TILE FLOORING"),
    (("marble",),                   "TILE FLOORING"),
    (("mosaic", "tile"),            "TILE FLOORING"),
    (("subway", "tile"),            "TILE FLOORING"),
    (("border", "tile"),            "TILE FLOORING"),
    (("terrazo",),                  "TILE FLOORING"),
    (("designer", "tile"),          "TILE FLOORING"),
    (("tile",),                     "TILE FLOORING"),
    (("carpet",),                   "TEXTILE FLOORING"),
    (("rug",),                      "TEXTILE FLOORING"),
    (("flocked", "flooring"),       "TEXTILE FLOORING"),
    (("raised", "floor"),           "RESILIENT FLOORING"),
    (("antistatic",),               "RESILIENT FLOORING"),
    (("spc",),                      "RESILIENT FLOORING"),
    (("lvt",),                      "RESILIENT FLOORING"),
    (("rubber", "flooring"),        "RESILIENT FLOORING"),
    (("linoleum",),                 "RESILIENT FLOORING"),
    (("laminated", "wooden"),       "RESILIENT FLOORING"),
    (("skirting",),                 "SKIRTING AND PROFILE"),
    (("transition", "profile"),     "SKIRTING AND PROFILE"),
    (("chair", "guard"),            "SKIRTING AND PROFILE"),
    (("corner", "guard"),           "SKIRTING AND PROFILE"),

    # ── CEILING ───────────────────────────────────────────────────────────
    (("gypsum", "ceiling"),         "GYPSUM CEILING"),
    (("laminated", "gypsum"),       "GYPSUM CEILING"),
    (("open", "cell"),              "METAL CEILING"),
    (("baffle", "ceiling"),         "METAL CEILING"),
    (("linear", "ceiling"),         "METAL CEILING"),
    (("metal", "ceiling"),          "METAL CEILING"),
    (("grid", "ceiling"),           "METAL CEILING"),
    (("wooden", "ceiling"),         "DECORATIVE CEILING"),
    (("stretch", "ceiling"),        "DECORATIVE CEILING"),
    (("sunlight", "ceiling"),       "DECORATIVE CEILING"),
    (("wooden", "baffle"),          "WOOD CEILING"),
    (("wooden", "open", "cell"),    "WOOD CEILING"),

    # ── PARTITIONS AND DOORS ──────────────────────────────────────────────
    (("glass", "partition"),        "GLASS PARTITION"),
    (("fixed", "glass"),            "GLASS PARTITION"),
    (("gypsum", "partition"),       "PARTITION"),
    (("acoustic", "partition"),     "PARTITION"),
    (("glass", "door"),             "DOORS"),
    (("wooden", "door"),            "DOORS"),
    (("panel", "door"),             "DOORS"),
    (("door", "handle"),            "DOORS"),
    (("door",),                     "DOORS"),
    (("aluminium", "window"),       "WINDOWS"),
    (("upvc", "window"),            "WINDOWS"),
    (("window",),                   "WINDOWS"),

    # ── DECORATIVES ───────────────────────────────────────────────────────
    (("roller", "blind"),           "BLIND"),
    (("honeycomb", "blind"),        "BLIND"),
    (("venetian", "blind"),         "BLIND"),
    (("zebra", "blind"),            "BLIND"),
    (("vertical", "blind"),         "BLIND"),
    (("blind",),                    "BLIND"),
    (("wallpaper",),                "GRAPHICS"),
    (("glass", "film"),             "GRAPHICS"),
    (("glass", "board"),            "GRAPHICS"),
    (("decal",),                    "GRAPHICS"),
    (("frame",),                    "GRAPHICS"),
    (("wall", "framing"),           "GRAPHICS"),
    (("pin", "up"),                 "GRAPHICS"),
    (("fluted", "panel"),           "PANELLING"),
    (("laminate", "panel"),         "PANELLING"),
    (("custom", "panel"),           "PANELLING"),
    (("3d", "mdf"),                 "PANELLING"),
    (("pvc", "louver"),             "PANELLING"),
    (("charcoal", "board"),         "PANELLING"),
    (("veneer", "panel"),           "PANELLING"),
    (("fluted",),                   "PANELLING"),
    (("panelling",),                "PANELLING"),
    (("wall", "panelling"),         "PANELLING"),
    (("writable",),                 "PANELLING"),
    (("creeper",),                  "PANELLING"),
    (("canvas",),                   "PANELLING"),
    (("wall", "cladding"),          "WALL CLADDING"),
    (("metal", "cladding"),         "WALL CLADDING"),
    (("branding",),                 "SIGNAGES AND BRANDING"),
    (("logo",),                     "SIGNAGES AND BRANDING"),
    (("signage",),                  "SIGNAGES AND BRANDING"),
    (("hardware",),                 "UTILITIES"),
    (("washroom", "mirror"),        "UTILITIES"),
    (("mirror",),                   "UTILITIES"),
    (("artifact",),                 "ARTIFACTS AND ACCESSORIES"),
    (("acoustic", "wall"),          "ACOUSTIC WALL SOLUTIONS"),
    (("movable", "wall"),           "ACOUSTIC WALL SOLUTIONS"),
    (("grooved", "panel"),          "ACOUSTIC WALL SOLUTIONS"),
    (("moulded", "acoustic"),       "ACOUSTIC WALL SOLUTIONS"),

    # ── SURFACE AND FINISHES ──────────────────────────────────────────────
    (("laminates",),                "WOODEN"),
    (("laminate",),                 "WOODEN"),
    (("veneer",),                   "WOODEN"),

    # ── CIVIL ─────────────────────────────────────────────────────────────
    (("water", "closet"),           "SANITARY FIXTURES"),
    (("wash", "basin"),             "SANITARY FIXTURES"),
    (("urinal",),                   "SANITARY FIXTURES"),
    (("faucet",),                   "SANITARY FIXTURES"),
    (("shower",),                   "SANITARY FIXTURES"),
    (("cistern",),                  "SANITARY FIXTURES"),
    (("soap", "dispenser"),         "SANITARY FIXTURES"),
    (("hand", "dryer"),             "SANITARY FIXTURES"),
    (("sanitary",),                 "SANITARY FIXTURES"),
    (("dado", "tile"),              "DADO TILE"),
    (("ceramic", "dado"),           "DADO TILE"),
    (("counter", "top"),            "COUNTER TOPS"),
    (("granite", "counter"),        "COUNTER TOPS"),
    (("quartz", "counter"),         "COUNTER TOPS"),
    (("fabricated", "structure"),   "FABRICATION"),
    (("fabricat",),                 "FABRICATION"),

    # ── PAINT ─────────────────────────────────────────────────────────────
    (("ceiling", "paint"),          "CEILING PAINT"),
    (("wall", "paint"),             "WALL PAINT"),
    (("duco", "paint"),             "DUCO PAINT"),
    (("duct", "paint"),             "DUCT PAINT"),
    (("texture", "paint"),          "TEXTURE PAINT"),

    # ── LIGHTING ──────────────────────────────────────────────────────────
    (("pendant", "light"),          "DECORATIVE LIGHTS"),
    (("chandelier",),               "DECORATIVE LIGHTS"),
    (("floor", "lamp"),             "DECORATIVE LIGHTS"),
    (("table", "lamp"),             "DECORATIVE LIGHTS"),
    (("wall", "sconce"),            "DECORATIVE LIGHTS"),
    (("wall", "light"),             "DECORATIVE LIGHTS"),
    (("led", "track", "light"),     "AMBIENT LIGHTS"),
    (("led", "downlight"),          "AMBIENT LIGHTS"),
    (("led", "panel", "light"),     "AMBIENT LIGHTS"),
    (("led", "strip"),              "AMBIENT LIGHTS"),
    (("led", "linear"),             "AMBIENT LIGHTS"),
    (("cob", "light"),              "AMBIENT LIGHTS"),
    (("profile", "light"),          "ARCHITECTURAL LIGHTS"),
    (("magnetic", "light"),         "ARCHITECTURAL LIGHTS"),
    (("acoustic", "linear", "light"), "ACOUSTIC LIGHTS"),
    (("acoustic", "pendant"),       "ACOUSTIC LIGHTS"),
    (("acoustic", "light"),         "ACOUSTIC LIGHTS"),

    # ── ACOUSTIC ──────────────────────────────────────────────────────────
    (("acoustic", "baffle"),        "CEILING"),
    (("acoustic", "tile"),          "CEILING"),
    (("moulded", "acoustic", "ceiling"), "CEILING"),
    (("privacy", "screen"),         "PARTITION"),
]


# ── Column alias detection ────────────────────────────────────────────────────

COLUMN_ALIASES = {
    "desc":     ("scope of work", "description", "particulars", "item description",
                 "scope", "work", "item", "details", "item name", "name"),
    "spec":     ("specification", "specifications", "spec", "technical spec",
                 "make & spec", "description & spec", "make and spec"),
    "make":     ("make", "brand", "make/brand", "make / brand", "manufacturer",
                 "make & brand", "make and brand", "makes"),
    "unit":     ("unit", "uom", "units", "unit of measurement", "u/m"),
    "qty":      ("qty", "quantity", "nos", "qnty", "design qty", "bcs qty",
                 "bcs quantity", "design quantity", "approved qty"),
    "site_qty": ("on site qty", "ops qty", "site qty", "onsite qty",
                 "site quantity", "on-site qty", "on site quantity"),
    "rate":     ("unit rate", "rate per unit", "rate", "bcs rate",
                 "unit rate (rs)", "unit rate(rs)", "u/rate", "ur"),
    "amount":   ("amount", "bcs amount", "total amount", "value",
                 "total value", "total", "bcs total"),
    "location": ("location", "space", "area", "space/area", "space / area",
                 "floor", "area / location"),
    "config":   ("configuration", "config", "size", "size/spec", "dimensions",
                 "dim", "size / spec"),
    "sr":       ("sr no", "sr. no", "s.no", "sno", "no", "item no", "sl no",
                 "sl. no", "s. no", "serial no"),
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").lower().strip())


def _cat_from_keywords(desc: str, default_cat: str = None) -> str:
    d = _norm(desc)
    for tokens, cat in KEYWORD_CATEGORY:
        if all(t in d for t in tokens):
            return cat
    return default_cat


def _subcat_from_keywords(desc: str) -> str | None:
    d = _norm(desc)
    for tokens, sub in KEYWORD_SUBCAT:
        if all(t in d for t in tokens):
            return sub
    return None


def _sheet_category(sheet_name: str) -> str | None:
    s = _norm(sheet_name)
    tokens = set(s.split())
    for rule_tokens, cat in SHEET_CATEGORY_TOKENS:
        if all(t in tokens or t in s for t in rule_tokens):
            return cat
    return None


def _detect_header_row(ws, max_scan: int = 12) -> int:
    """Find the row most likely to be the header by scoring alias token hits."""
    all_tokens: set[str] = set()
    for aliases in COLUMN_ALIASES.values():
        for alias in aliases:
            all_tokens.update(alias.split())

    best_row, best_score = 1, 0
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        score = 0
        for c in range(1, min(ws.max_column + 1, 25)):
            cell = _norm(str(ws.cell(r, c).value or ""))
            score += len(set(cell.split()) & all_tokens)
        if score > best_score:
            best_score, best_row = score, r
    return best_row


def _map_columns(ws, header_row: int) -> dict:
    """Return {field: col_index (1-based)} by matching headers against COLUMN_ALIASES."""
    result = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(str(ws.cell(header_row, c).value or ""))
        if not h:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if field in result:
                continue
            if any(alias == h or alias in h or h in alias for alias in aliases):
                result[field] = c
                break
    return result


def _try_float(v) -> float | None:
    try:
        return float(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _make_item(boq_no, category, subcat, desc, spec, make, cfg, uom,
               qty, rate, amount, location, section_label) -> dict:
    return {
        "BOQ No":             boq_no,
        "Section":            section_label,
        "Category":           category or "UNCATEGORIZED",
        "Sub-category":       subcat or "",
        "Item":               "",
        "Description":        _clean(desc),
        "Specification":      _clean(spec),
        "Make/Brand":         _clean(make),
        "Config (size/spec)": _clean(cfg or spec)[:80],
        "UoM":                str(uom or "").strip().upper(),
        "BCS Qty":            qty,
        "BCS Rate":           rate,
        "BCS Amount":         amount,
        "Space/Area":         _clean(location),
    }


# ── Format detection ──────────────────────────────────────────────────────────

_SINGLE_SHEET_NAMES = {
    "boq", "budget", "smart budget", "bill of quantities",
    "quotation", "quote", "items", "boq sheet", "item list",
}


def _detect_single_sheet(wb) -> str | None:
    """Return sheet name if this looks like a single-sheet BOQ, else None."""
    for s in wb.sheetnames:
        if _norm(s) in _SINGLE_SHEET_NAMES:
            return s
    if len(wb.sheetnames) == 1:
        return wb.sheetnames[0]
    return None


# ── Single-sheet loader ───────────────────────────────────────────────────────

def _load_single_sheet(ws) -> list[dict]:
    header_row = _detect_header_row(ws)
    col = _map_columns(ws, header_row)

    desc_c   = col.get("desc", 2)
    spec_c   = col.get("spec", 3)
    make_c   = col.get("make", 4)
    unit_c   = col.get("unit", 5)
    qty_c    = col.get("qty", 6)
    rate_c   = col.get("rate", 7)
    amount_c = col.get("amount", 8)
    loc_c    = col.get("location")
    cfg_c    = col.get("config")

    items = []
    current_section_code = None
    current_section_name = None
    item_seq = 0

    for r in range(header_row + 1, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        desc  = ws.cell(r, desc_c).value
        qty   = ws.cell(r, qty_c).value   if qty_c   else None
        rate  = ws.cell(r, rate_c).value  if rate_c  else None
        amount= ws.cell(r, amount_c).value if amount_c else None
        spec  = ws.cell(r, spec_c).value  if spec_c  else None
        make  = ws.cell(r, make_c).value  if make_c  else None
        unit  = ws.cell(r, unit_c).value  if unit_c  else None
        loc   = ws.cell(r, loc_c).value   if loc_c   else None
        cfg   = ws.cell(r, cfg_c).value   if cfg_c   else None

        if all(v in (None, "") for v in (col_a, desc, qty, rate)):
            continue

        desc_str = _clean(desc)

        # Skip total rows
        if desc_str and any(t in desc_str.upper() for t in
                            ("SUB TOTAL", "GRAND TOTAL", " TOTAL")):
            if not qty and not rate:
                continue

        # Section header: single uppercase letter in col A, no qty/rate
        col_a_str = str(col_a or "").strip()
        if len(col_a_str) == 1 and col_a_str.isupper() and not qty and not rate:
            current_section_code = col_a_str
            current_section_name = desc_str
            continue

        if not desc_str or (not qty and not rate):
            continue

        item_seq += 1
        boq_no = f"{current_section_code or 'X'}{item_seq:03d}"
        default_cat = SECTION_CODE_CATEGORY.get(current_section_code)
        cat = _cat_from_keywords(desc_str, default_cat)
        sub = _subcat_from_keywords(desc_str)
        section_label = f"{current_section_code or '?'} | {current_section_name or ''}"

        items.append(_make_item(
            boq_no, cat, sub, desc_str, spec, make, cfg, unit,
            _try_float(qty), _try_float(rate), _try_float(amount),
            loc, section_label,
        ))

    return items


# ── Multi-sheet loader ────────────────────────────────────────────────────────

def _load_multi_sheet(wb) -> list[dict]:
    items = []
    seq = 0

    for sheet_name in wb.sheetnames:
        sn = _norm(sheet_name)
        if any(t in sn for t in SKIP_SHEET_TOKENS):
            continue

        category = _sheet_category(sheet_name)
        if category is None:
            continue

        ws = wb[sheet_name]
        header_row = _detect_header_row(ws)
        col = _map_columns(ws, header_row)

        sr_c      = col.get("sr", 1)
        desc_c    = col.get("desc", 2)
        loc_c     = col.get("location", 3)
        spec_c    = col.get("spec", 4)
        cfg_c     = col.get("config", 5)
        make_c    = col.get("make", 6)
        unit_c    = col.get("unit", 8)
        qty_c     = col.get("qty", 9)
        siteqty_c = col.get("site_qty", 10)
        rate_c    = col.get("rate", 12)
        amount_c  = col.get("amount", 13)

        current_subcat = ""

        for r in range(header_row + 1, ws.max_row + 1):
            sr   = ws.cell(r, sr_c).value
            desc = ws.cell(r, desc_c).value
            qty  = ws.cell(r, qty_c).value  if qty_c  else None
            rate = ws.cell(r, rate_c).value if rate_c else None

            if not desc and not qty and not rate:
                continue

            desc_str = _clean(desc)
            sr_str   = str(sr or "").strip()

            # Skip total rows
            if "total" in desc_str.lower() and not qty and not rate:
                continue

            # Sub-section header: letter+digit code (G1, C2…), no qty/rate
            if re.match(r"^[A-Z]\d+$", sr_str) and not qty and not rate:
                current_subcat = SUBSECTION_CODE_SUBCAT.get(sr_str, desc_str[:30])
                continue

            # Must look like a numeric item row
            if not re.match(r"^\d+(\.\d+)?$", sr_str):
                continue
            if not desc_str:
                continue

            spec  = ws.cell(r, spec_c).value   if spec_c   else None
            make  = ws.cell(r, make_c).value   if make_c   else None
            unit  = ws.cell(r, unit_c).value   if unit_c   else None
            cfg   = ws.cell(r, cfg_c).value    if cfg_c    else None
            loc   = ws.cell(r, loc_c).value    if loc_c    else None
            s_qty = ws.cell(r, siteqty_c).value if siteqty_c else None
            amt   = ws.cell(r, amount_c).value  if amount_c  else None

            final_qty = _try_float(s_qty) or _try_float(qty)
            cat = _cat_from_keywords(desc_str, category)
            sub = _subcat_from_keywords(desc_str) or current_subcat

            seq += 1
            items.append(_make_item(
                f"{sheet_name[:3].upper()}{seq:04d}",
                cat, sub, desc_str, spec, make, cfg, unit,
                final_qty, _try_float(rate), _try_float(amt),
                loc, f"{sheet_name} / {current_subcat}",
            ))

    return items


# ── Public API ────────────────────────────────────────────────────────────────

def load_boq(path: str) -> list[dict]:
    """
    Load any BOQ Excel file, auto-detecting its structure.
    Returns list of dicts with keys:
      BOQ No, Section, Category, Sub-category, Item, Description,
      Specification, Make/Brand, Config (size/spec), UoM,
      BCS Qty, BCS Rate, BCS Amount, Space/Area
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    single = _detect_single_sheet(wb)
    if single:
        return _load_single_sheet(wb[single])
    return _load_multi_sheet(wb)
