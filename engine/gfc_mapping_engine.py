"""
GFC ↔ Smart Budget Mapping — Orchestrator

Implements the two-engine pipeline:

  ENGINE 1 (gfc_classifier.py)
    Maps every GFC line item to Category + Sub-category + Item
    from the Category Master before any BOQ matching runs.
    Goal: near-100 % taxonomy accuracy.

  ENGINE 2 (boq_matcher.py)
    Matches each classified GFC item against the BOQ/Smart Budget
    using an 11-signal scorer.  Returns 🟢 Auto / 🟡 Suggested / 🔴 Not in BOQ.

Layers still handled here:
  L1  Sheet name → Category        (classify_sheet, SHEET_RULES)
  L2  Column headers → field map   (map_columns, detect_header_row)
  L3  Row classification            (classify_row)
  L4  Field normalization           (normalize_status, normalize_uom, parse_bcs_rate)
  L7  Multi-floor explosion         (explode_multi_floor)
  Loaders  load_category_master
  Orchestrator  process_gfc
"""
import openpyxl
import re
from collections import defaultdict
from dataclasses import dataclass
from typing import Optional

# Engine 1 and Engine 2 — imported here, used in process_gfc
from .gfc_classifier import ClassificationResult, classify_line_item
from .boq_matcher import best_match, MATCH_THRESHOLDS   # re-exported for gui + output_builder
from . import image_classifier as _img_mod

__all__ = [
    "classify_sheet", "SHEET_RULES", "SKIP_TOKENS",
    "map_columns", "detect_header_row", "FIELD_ALIASES",
    "classify_row",
    "normalize_status", "normalize_uom", "parse_bcs_rate", "parse_indian_number",
    "load_category_master",
    "EnrichedLine", "process_gfc",
    "MATCH_THRESHOLDS",   # keep importable from here for backward compatibility
]


# ============================================================
# L1 — SHEET NAME → CATEGORY
# ============================================================
SHEET_RULES = [
    # (tokens that must appear, Category, Sub-category hint)
    # Sub-category hints use exact Category Master names.

    # ── FURNITURE ─────────────────────────────────────────────────────────
    (("modular",),                  "FURNITURE",            "MODULAR FURNITURE"),
    (("loose", "furniture"),        "FURNITURE",            "LOOSE FURNITURE"),
    (("chair",),                    "FURNITURE",            "CHAIRS"),
    (("carpentry",),                "FURNITURE",            "CUSTOM FURNITURE"),
    (("carpentary",),               "FURNITURE",            "CUSTOM FURNITURE"),
    (("caprentry",),                "FURNITURE",            "CUSTOM FURNITURE"),
    (("carpentery",),               "FURNITURE",            "CUSTOM FURNITURE"),
    (("furniture", "list"),         "FURNITURE",            None),
    (("rack",),                     "FURNITURE",            "MODULAR FURNITURE"),

    # ── FLOORING ──────────────────────────────────────────────────────────
    (("tile",),                     "FLOORING",             "TILE FLOORING"),
    (("carpet",),                   "FLOORING",             "TEXTILE FLOORING"),
    (("spc",),                      "FLOORING",             "RESILIENT FLOORING"),
    (("lvt",),                      "FLOORING",             "RESILIENT FLOORING"),
    (("vinyl",),                    "FLOORING",             "RESILIENT FLOORING"),
    (("antistatic",),               "FLOORING",             "RESILIENT FLOORING"),
    (("anti", "static"),            "FLOORING",             "RESILIENT FLOORING"),
    (("flooring",),                 "FLOORING",             None),
    # Skirting/transition → DECORATIVES per Category Master
    (("skirting",),                 "DECORATIVES",          "SKIRTING AND PROFILE"),
    (("transition",),               "DECORATIVES",          "SKIRTING AND PROFILE"),

    # ── CEILING ───────────────────────────────────────────────────────────
    (("open", "cell"),              "CEILING",              "METAL CEILING"),
    (("baffle", "ceiling"),         "CEILING",              "WOOD CEILING"),  # baffles default to wooden
    (("ceiling",),                  "CEILING",              None),

    # ── PAINT ─────────────────────────────────────────────────────────────
    (("paint",),                    "PAINT",                None),

    # ── ACOUSTIC ──────────────────────────────────────────────────────────
    # "Acoustic Panel" sheets = ACOUSTIC category (functional sound absorption).
    # Decorative panels with acoustic properties on wall/panelling sheets stay DECORATIVES.
    (("acoustic", "panel"),         "ACOUSTIC",             None),
    (("acoustic", "wall"),          "ACOUSTIC",             "WALL SOLUTIONS"),
    (("accoustic", "panel"),        "ACOUSTIC",             None),
    (("acoustic",),                 "ACOUSTIC",             None),
    (("accoustic",),                "ACOUSTIC",             None),

    # ── LIGHTING ──────────────────────────────────────────────────────────
    (("light",),                    "LIGHTING",             None),
    (("electrical",),               "LIGHTING",             None),

    # ── PARTITIONS AND DOORS ──────────────────────────────────────────────
    (("phenolic",),                 "PARTITIONS AND DOORS", "PARTITION"),
    (("partition",),                "PARTITIONS AND DOORS", "PARTITION"),
    (("door",),                     "PARTITIONS AND DOORS", "DOORS"),
    (("glazing",),                  "PARTITIONS AND DOORS", "GLASS PARTITION"),

    # ── DECORATIVES ───────────────────────────────────────────────────────
    (("blind", "roller"),           "DECORATIVES",          "BLIND"),
    (("blind",),                    "DECORATIVES",          "BLIND"),
    (("wallpaper",),                "DECORATIVES",          "GRAPHICS"),
    (("fluted",),                   "DECORATIVES",          "PANELLING"),
    (("back", "painted", "glass"),  "DECORATIVES",          "GRAPHICS"),
    (("glass", "film"),             "DECORATIVES",          "GRAPHICS"),
    (("pin", "up"),                 "DECORATIVES",          "GRAPHICS"),
    (("logo",),                     "DECORATIVES",          "SIGNAGES AND BRANDING"),
    (("signage",),                  "DECORATIVES",          "SIGNAGES AND BRANDING"),
    (("branding",),                 "DECORATIVES",          "SIGNAGES AND BRANDING"),
    (("frame",),                    "DECORATIVES",          "GRAPHICS"),
    (("planter",),                  "DECORATIVES",          None),
    (("creeper",),                  "DECORATIVES",          "PANELLING"),
    (("writable",),                 "DECORATIVES",          "PANELLING"),
    (("decal",),                    "DECORATIVES",          "GRAPHICS"),
    (("fabric",),                   "DECORATIVES",          "PANELLING"),
    # "canvas frame" = framed canvas artwork = GRAPHICS. Must be before generic "canvas".
    (("canvas", "frame"),           "DECORATIVES",          "GRAPHICS"),
    (("canvas",),                   "DECORATIVES",          "PANELLING"),
    (("panelling",),                "DECORATIVES",          "PANELLING"),
    (("cladding",),                 "DECORATIVES",          "WALL CLADDING"),

    # ── SURFACE AND FINISHES ──────────────────────────────────────────────
    (("laminate",),                 "SURFACE AND FINISHES", "WOODEN"),
    (("veneer",),                   "SURFACE AND FINISHES", "WOODEN"),
    # Corian / HI-MACS / solid surface = surface finish material → WOODEN sub-cat
    (("corian",),                   "SURFACE AND FINISHES", "WOODEN"),
    (("hi", "macs"),                "SURFACE AND FINISHES", "WOODEN"),
    (("solid", "surface"),          "SURFACE AND FINISHES", "WOODEN"),

    # ── CIVIL ─────────────────────────────────────────────────────────────
    (("ms", "fabricat"),            "CIVIL",                "FABRICATION"),
    (("ms", "fab"),                 "CIVIL",                "FABRICATION"),
    (("fabricat",),                 "CIVIL",                "FABRICATION"),
    (("sanitary",),                 "CIVIL",                "SANITARY FIXTURES"),
    (("dado",),                     "CIVIL",                "DADO TILE"),
    (("counter",),                  "CIVIL",                "COUNTER TOPS"),
]

SKIP_TOKENS = ("drawing schedule", "mom", "to do", "todo", "cumulative",
               "timeline", "3ds plan", "legend", "light dump", "sheet1",
               "tender", "tv size", "index")


def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9 ]", " ", (s or "").lower()).strip()


def classify_sheet(sheet_name: str) -> tuple[Optional[str], Optional[str], float, str]:
    """L1 — Returns (category, sub_hint, confidence 0-1, reason)."""
    n = normalize(sheet_name)
    for tok in SKIP_TOKENS:
        if tok in n:
            return (None, None, 1.0, f"SKIP: '{tok}' matched")
    best = None
    for tokens, cat, sub in SHEET_RULES:
        if all(t in n for t in tokens):
            score = sum(len(t) for t in tokens) / max(len(n), 1)
            if best is None or score > best[2]:
                best = (cat, sub, score, f"matched tokens {tokens}")
    return best if best else (None, None, 0.0, "no rule matched")


# ============================================================
# L2 — COLUMN HEADERS → CANONICAL FIELDS
# ============================================================
FIELD_ALIASES = {
    "sr_no":        {"sr no", "sr. no", "sr.no", "s.no", "sno", "sl no", "sl. no",
                     "a", "s. no.", "sr. no.", "wall no.", "wall no", "re"},
    "description":  {"description", "description - scope of work", "scope of work",
                     "item", "tag", "furniture/props/paneling", "application",
                     "discription", "product", "details", "details (dimension drawing)",
                     "description - scope of work", "item of work",
                     "scope", "work description", "particulars"},
    "subcat_hint":  {"furniture type",
                     "wall/ ceiling / slab", "wall ceiling slab",
                     "wall/ceiling/slab", "surface type", "surface"},
    "areas":        {"areas", "area", "location", "room", "space",
                     "location as per layout/on-site",
                     "location/area", "location as per layout on-site"},
    "size":         {"size", "size in inch & feet", "size (l x d x h)", "dimension",
                     "dimensions", "size in inch", "size in feet", "dia/size",
                     "size (mm)", "dimensions (wxh)", "size in mm"},
    "size_l":       {"length", "l"},
    "size_w":       {"width", "w"},
    "size_h":       {"height", "h"},
    "uom":          {"unit", "uom", "u.o.m"},
    "qty_design":   {"qty", "quantity", "design qty", "design quantity",
                     "qty (design)", "qty design", "design qty.", "qty.",
                     "required qty", "quantity sqft + 10% wastage",
                     "qty in nos", "area (sqft)", "sqft area", "design quantity"},
    "qty_ops":      {"ops qty", "site qty", "on site", "on site qty",
                     "qty (ops)", "actual qty", "ops quantity", "site",
                     "ops qty (final)", "ops qty"},
    "brand":        {"brand", "make", "manufacturer", "make - brand",
                     "catalogue/ brand name", "brand name", "makes"},
    "catalogue":    {"catalogue name", "catalogue", "catelogue name",
                     "catalogue/code", "tile name"},
    "product_name": {"product name", "paint name", "tile image"},
    "finish":       {"finish"},
    "finish_code":  {"code", "product code", "cat no", "cat no.", "laminate code",
                     "finishes (code)", "ral shade", "color code", "code planter"},
    "rate":         {"rate", "bcs rate", "approx rate", "rate /unit", "rate per unit",
                     "boq rate", "bcs"},
    "status":       {"status", "approval status", "client status"},
    "ref_image":    {"render image", "image", "reference image", "ref image",
                     "product image", "3d render", "3d references",
                     "approved refrence 3d view"},
}

_NORMALIZED_ALIASES = {
    fld: {re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", a.lower()).strip()).strip()
          for a in als}
    for fld, als in FIELD_ALIASES.items()
}


def map_columns(header_row: list) -> dict:
    """L2 — {canonical_field: column_index}."""
    mapping = {}
    for col_idx, val in enumerate(header_row):
        if val is None:
            continue
        norm = normalize(str(val))
        for field_name, aliases in _NORMALIZED_ALIASES.items():
            if norm in aliases:
                if field_name not in mapping:
                    mapping[field_name] = col_idx
                break
            matched = False
            for alias in aliases:
                if alias in norm and len(alias) > 4:
                    if field_name not in mapping:
                        mapping[field_name] = col_idx
                    matched = True
                    break
            if matched:
                break
    return mapping


def detect_header_row(rows: list, max_scan: int = 12) -> tuple[Optional[int], dict]:
    """Scan first max_scan rows; return (row_index, column_map) for best header."""
    best_idx, best_map, best_score = None, {}, 0
    for i, row in enumerate(rows[:max_scan]):
        m = map_columns(row)
        score = len(m)
        if "description" in m or "areas" in m or "product_name" in m:
            score += 2
        if "qty_design" in m or "qty_ops" in m:
            score += 1
        if "uom" in m:
            score += 1
        if score > best_score:
            best_score, best_idx, best_map = score, i, m
    return (best_idx, best_map)


# ============================================================
# L3 — ROW CLASSIFICATION
# ============================================================
SECTION_HEADER_PATTERNS = [
    re.compile(r"^\s*\d+\s*(st|nd|rd|th)?\s*floor", re.I),
    re.compile(r"^(flooring|modular|lighting|cabin|workstation|cafeteria|meeting)\s*[:]*\s*$", re.I),
    re.compile(r"^[A-Z\s]+:$"),
    re.compile(r"^option\s*\d+", re.I),
]
TOTAL_PATTERNS = [
    re.compile(r"^\s*total", re.I),
    re.compile(r"^\s*grand total", re.I),
    re.compile(r"^\s*sub\s*total", re.I),
]
EXISTING_REPLACE = re.compile(r"existing.*(to be|replace)", re.I)


def classify_row(row: list, col_map: dict) -> str:
    """L3 — item / subrow / section / total / note / empty."""
    non_empty = [c for c in row if c not in (None, "")]
    if not non_empty:
        return "empty"

    desc_val = None
    for cand in ("description", "product_name", "areas"):
        idx = col_map.get(cand)
        if idx is not None and idx < len(row) and row[idx] not in (None, ""):
            desc_val = row[idx]
            break
    desc_str = str(desc_val).strip() if desc_val is not None else ""

    if len(non_empty) == 1 and isinstance(non_empty[0], str):
        s = non_empty[0]
        for p in SECTION_HEADER_PATTERNS:
            if p.search(s):
                return "section"
        if EXISTING_REPLACE.search(s):
            return "section"
        if len(s) < 60 and s.isupper():
            return "section"
        return "note"

    for c in row:
        if c and isinstance(c, str):
            for p in TOTAL_PATTERNS:
                if p.match(c.strip()):
                    return "total"

    sr_idx  = col_map.get("sr_no")
    has_sr  = sr_idx is not None and sr_idx < len(row) and row[sr_idx] not in (None, "")
    has_qty = any(
        col_map.get(k) is not None
        and col_map[k] < len(row)
        and row[col_map[k]] not in (None, "")
        for k in ("qty_design", "qty_ops")
    )
    if not has_sr and desc_str.strip() and has_qty:
        return "subrow"
    if desc_str.strip() and (has_qty or has_sr):
        return "item"
    return "note"


# ============================================================
# L4 — FIELD NORMALIZATION
# ============================================================
STATUS_RULES = {
    "approved":                       ("Client Approved", "exact match"),
    "approved by client":             ("Client Approved", "exact match"),
    "client approved":                ("Client Approved", "exact match"),
    "closed":                         ("Client Approved", "treated as approved"),
    "complete":                       ("Client Approved", "treated as approved"),
    "completed":                      ("Client Approved", "treated as approved"),
    "done":                           ("Client Approved", "treated as approved"),
    "ok":                             ("Client Approved", "shorthand"),
    "ordered":                        ("Client Approved", "post-approval"),
    "order done":                     ("Client Approved", "post-approval"),
    "pending":                        ("Pending", "exact"),
    "wip":                            ("Pending", "work in progress"),
    "open":                           ("Pending", "treated as pending"),
    "approval pending":               ("Pending", "explicit"),
    "in process":                     ("Pending", "treated as pending"),
    "removed":                        ("Excluded", "removed"),
    "cancelled":                      ("Excluded", "cancelled"),
    "canceled":                       ("Excluded", "cancelled"),
    "na":                             ("Excluded", "n/a"),
    "n/a":                            ("Excluded", "n/a"),
    "-":                              ("Excluded", "dash"),
    "not to be considered for order": ("Excluded", "explicit"),
}


def normalize_status(raw) -> tuple[str, str]:
    if raw is None or str(raw).strip() == "":
        return ("Pending", "blank → Pending")
    s = str(raw).strip().lower()
    if s in STATUS_RULES:
        return STATUS_RULES[s]
    if "removed" in s or "cancel" in s or "not to be" in s:
        return ("Excluded", f"exclusion keyword: '{s}'")
    if any(k in s for k in ("approv", "done", "closed", "complete",
                             "ordered", "received")) and "pend" not in s:
        return ("Client Approved", f"approval keyword: '{s}'")
    if "pend" in s or "wip" in s or "open" in s:
        return ("Pending", f"pending keyword: '{s}'")
    if re.search(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", s):
        return ("Pending", f"date stamp only → Pending")
    return ("Pending", f"unrecognized '{s}' → Pending")


UOM_MAP = {
    "nos": "NOS", "nos.": "NOS", "no": "NOS", "no.": "NOS", "pcs": "NOS", "pc": "NOS",
    "qty": "NOS", "unit": "NOS", "set": "NOS", "each": "NOS",
    "sqft": "SQFT", "sq ft": "SQFT", "sq.ft": "SQFT", "sft": "SQFT", "sq feet": "SQFT",
    "sqm": "SQM", "sqmt": "SQM", "sq m": "SQM", "sq.m": "SQM", "sq meter": "SQM",
    "rft": "RFT", "rmt": "RMT", "lft": "RFT", "l.ft": "RFT", "running ft": "RFT",
    "kg": "KG", "ton": "TON", "litre": "LITRE", "l": "LITRE",
}


def normalize_uom(raw) -> tuple[Optional[str], str]:
    if raw is None or str(raw).strip() == "":
        return (None, "blank")
    s = str(raw).strip().lower().rstrip(".")
    if s in UOM_MAP:
        return (UOM_MAP[s], f"'{raw}' → {UOM_MAP[s]}")
    for k, v in UOM_MAP.items():
        if k in s or s in k:
            return (v, f"fuzzy '{raw}' → {v}")
    return (str(raw).upper(), f"unrecognized '{raw}'")


def parse_indian_number(s: str) -> Optional[float]:
    if s is None:
        return None
    s = str(s).strip().lower().replace(",", "").replace("rs", "").replace("₹", "").strip()
    if not s:
        return None
    m = re.match(r"([\d.]+)\s*cr", s)
    if m:
        return float(m.group(1)) * 1_00_00_000
    m = re.match(r"([\d.]+)\s*l", s)
    if m:
        return float(m.group(1)) * 1_00_000
    m = re.match(r"([\d.]+)", s)
    if m:
        return float(m.group(1))
    return None


def parse_bcs_rate(raw) -> tuple[Optional[float], Optional[str], str]:
    if raw is None or str(raw).strip() == "":
        return (None, None, "blank")
    s = str(raw).strip()
    per = None
    s_lower = s.lower()
    for alias in sorted(UOM_MAP.keys(), key=len, reverse=True):
        if re.search(r"\b" + re.escape(alias) + r"\b", s_lower):
            per = UOM_MAP[alias]
            break
    cleaned = re.sub(
        r"\b(per|rs|inr|/|nos?\.?|sqft|sq\.?ft|sqm|rft|rmt|lft|kg|ton)\b", " ", s_lower)
    val = parse_indian_number(cleaned)
    if val is not None:
        return (val, per, f"parsed '{raw}' → {val} per {per or '?'}")
    return (None, None, f"could not parse '{raw}'")


# ============================================================
# L7 — EDGE CASE HANDLERS
# ============================================================
FLOOR_COL_PATTERN = re.compile(r"^\s*\d{1,2}\s*(st|nd|rd|th)?\s*floor\s*$", re.I)


def explode_multi_floor(row: list, col_map: dict, header_row: list) -> list[tuple[list, str]]:
    floor_cols = [
        (i, str(h).strip())
        for i, h in enumerate(header_row)
        if h and FLOOR_COL_PATTERN.match(str(h))
    ]
    if not floor_cols:
        return [(row, "")]
    qty_col  = col_map.get("qty_ops") or col_map.get("qty_design")
    exploded = []
    for (idx, lbl) in floor_cols:
        if idx < len(row) and row[idx] not in (None, "", 0):
            new = list(row)
            if qty_col is not None and qty_col < len(new):
                new[qty_col] = row[idx]
            exploded.append((new, lbl))
    return exploded if exploded else [(row, "")]


# ============================================================
# LOADER
# ============================================================

def load_category_master(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        cat, sub, item, itype = (ws.cell(r, c).value for c in range(1, 5))
        if cat:
            rows.append({
                "Category":     cat,
                "Sub-category": sub or "",
                "Item":         item or "",
                "Item Type":    itype or "",
            })
    return rows


# ============================================================
# ENRICHED LINE — result record
# ============================================================

@dataclass
class EnrichedLine:
    # Provenance
    gfc_sheet:           str = ""
    gfc_row_idx:         int = 0
    floor_label:         str = ""

    # L1 — sheet classification
    detected_category:   Optional[str] = None
    detected_sub_hint:   Optional[str] = None
    sheet_match_reason:  str = ""
    row_type:            str = "item"

    # Raw fields + L4 normalized
    raw_description:     str = ""
    raw_brand:           str = ""
    raw_uom:             str = ""
    raw_qty_design:      Optional[float] = None
    raw_qty_ops:         Optional[float] = None
    raw_status:          str = ""
    raw_finish_code:     str = ""
    raw_areas:           str = ""
    raw_size:            str = ""
    norm_uom:            Optional[str] = None
    norm_status:         str = ""
    status_reason:       str = ""

    # Engine 1 — Category Master classification
    master_category:          Optional[str] = None
    master_subcategory:       Optional[str] = None
    master_item:              Optional[str] = None
    master_reason:            str = ""
    classification_confidence: int = 0      # 0–100
    classification_method:    str = ""      # keyword | item_match | sheet_hint | row_hint
    classification_signals:   str = ""      # pipe-separated audit trail

    # Engine 2 — BOQ match
    match_status:        str = ""
    match_score:         int = 0
    matched_boq_no:      Optional[str] = None
    matched_description: Optional[str] = None
    matched_brand:       Optional[str] = None
    matched_uom:         Optional[str] = None
    matched_bcs_rate:    Optional[float] = None
    score_breakdown:     str = ""

    # PR readiness
    pr_ready:            bool = False
    pr_blocker:          str = ""

    # Vision enrichment (Signal 5) — populated when vision_enabled=True and confidence < 75
    vision_subcategory:  Optional[str] = None
    vision_confidence:   int = 0
    vision_reason:       str = ""


# ============================================================
# ORCHESTRATOR
# ============================================================

def process_gfc(
    gfc_path: str,
    budget: list[dict],
    category_master: list[dict],
    limit_sheets: Optional[int] = None,
    run_match: bool = True,
    vision_enabled: bool = False,
    vision_api_key: Optional[str] = None,
) -> list[EnrichedLine]:
    """
    Two-engine pipeline:
      Pass 1 (Engine 1): classify every GFC item → Category Master taxonomy
      Pass 2 (Engine 2): match every classified item → best BOQ item + score

    run_match=False runs ENGINE 1 ONLY (category/sub-category mapping) and skips
    the BOQ matching pass — used when no BOQ/Smart Budget is supplied.
    """
    # ── Pre-index by category ─────────────────────────────────────────────
    budget_by_cat: dict[str, list] = defaultdict(list)
    for item in budget:
        budget_by_cat[(item.get("Category") or "").upper()].append(item)

    master_by_cat: dict[str, list] = defaultdict(list)
    for m in category_master:
        master_by_cat[(m.get("Category") or "").upper()].append(m)

    wb      = openpyxl.load_workbook(gfc_path, data_only=True, read_only=True)
    results: list[EnrichedLine] = []

    # Vision: extract all sheet images upfront (fast zip parse, no PIL decoding yet)
    _all_images: dict[str, dict[int, tuple]] = {}
    _vision_calls = 0
    if vision_enabled:
        _all_images = _img_mod.extract_all_images(gfc_path, list(wb.sheetnames))

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        if limit_sheets and sheet_idx >= limit_sheets:
            break

        # L1 — sheet → category
        cat, sub_hint, _conf, sheet_reason = classify_sheet(sheet_name)
        if cat is None:
            continue

        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]  # no row cap

        # L2 — header detection + column map
        header_idx, col_map = detect_header_row(rows)
        if header_idx is None or (
            "description" not in col_map
            and "product_name" not in col_map
            and "areas" not in col_map
        ):
            continue
        header_row = rows[header_idx]

        # Resolve pre-filtered slices for this sheet's category
        cat_upper  = cat.upper()
        cat_budget = budget_by_cat.get(cat_upper, [])
        cat_exists = len(cat_budget) > 0
        cat_master = master_by_cat.get(cat_upper, [])

        # ── Row loop ──────────────────────────────────────────────────────
        for r_offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if all(c in (None, "") for c in row):
                continue

            for (erow, floor_lbl) in explode_multi_floor(row, col_map, header_row):
                row_type = classify_row(erow, col_map)
                if row_type in ("empty", "total", "note", "section"):
                    continue

                el = EnrichedLine(
                    gfc_sheet          = sheet_name,
                    gfc_row_idx        = r_offset,
                    floor_label        = floor_lbl,
                    detected_category  = cat,
                    detected_sub_hint  = sub_hint,
                    sheet_match_reason = sheet_reason,
                    row_type           = row_type,
                )

                def cell(field, _r=erow, _m=col_map):
                    i = _m.get(field)
                    if i is None or i >= len(_r):
                        return ""
                    v = _r[i]
                    return "" if v is None else str(v).strip()

                el.raw_description = (cell("description")
                                      or cell("product_name")
                                      or cell("areas"))
                el.raw_brand       = cell("brand") or cell("catalogue")
                el.raw_uom         = cell("uom")
                el.raw_status      = cell("status")
                el.raw_finish_code = cell("finish_code")
                el.raw_areas       = cell("areas")

                sz = cell("size")
                if not sz:
                    parts = [p for p in (cell("size_l"), cell("size_w"), cell("size_h")) if p]
                    if parts:
                        sz = " x ".join(parts) + " mm"
                el.raw_size = sz

                # Row-level sub-cat hint (overrides sheet-level for L1 signal)
                row_subcat_raw = cell("subcat_hint")
                local_sub_hint = sub_hint
                if row_subcat_raw:
                    rs = row_subcat_raw.upper()
                    if   "CARPENT" in rs: local_sub_hint = "CUSTOM FURNITURE"
                    elif "MODULAR" in rs: local_sub_hint = "MODULAR FURNITURE"
                    elif "LOOSE"   in rs: local_sub_hint = "LOOSE FURNITURE"
                    elif "CHAIR"   in rs: local_sub_hint = "CHAIRS"
                    else:                 local_sub_hint = rs
                el.detected_sub_hint = local_sub_hint

                for attr, fld in (("raw_qty_design", "qty_design"),
                                  ("raw_qty_ops",    "qty_ops")):
                    v = cell(fld)
                    if v:
                        try:
                            setattr(el, attr, float(v))
                        except ValueError:
                            pass

                # L4 — normalize
                el.norm_uom, _ = normalize_uom(el.raw_uom)
                if not el.norm_uom:
                    el.norm_uom = {
                        "FURNITURE":            "NOS",
                        "LIGHTING":             "NOS",
                        "FLOORING":             "SQFT",
                        "PAINT":                "SQFT",
                        "ACOUSTIC":             "SQFT",
                        "DECORATIVES":          "SQFT",
                        "CEILING":              "SQFT",
                        "SURFACE AND FINISHES": "RFT",
                    }.get(cat)
                el.norm_status, el.status_reason = normalize_status(el.raw_status)

                # ── ENGINE 1: classify against Category Master ─────────────
                clf: ClassificationResult = classify_line_item(
                    category            = el.detected_category,
                    sub_hint            = el.detected_sub_hint,
                    description         = el.raw_description,
                    areas               = el.raw_areas,
                    product_name        = cell("product_name"),
                    row_subcat_hint     = row_subcat_raw,
                    master_for_category = cat_master,
                    master_by_cat       = master_by_cat,   # enables category override
                )
                el.master_category          = clf.category
                el.master_subcategory       = clf.subcategory
                el.master_item              = clf.item
                el.master_reason            = clf.method
                el.classification_confidence = clf.confidence
                el.classification_method    = clf.method
                el.classification_signals   = " | ".join(clf.signals)

                # ── VISION (Signal 5): image-based sub-cat refinement ─────────
                # Triggered only for MEDIUM/LOW confidence and when an image is
                # available for this row.  Hard-capped at MAX_IMAGE_CALLS per run.
                if (vision_enabled
                        and el.master_category
                        and clf.confidence < _img_mod.IMAGE_THRESHOLD
                        and _vision_calls < _img_mod.MAX_IMAGE_CALLS):
                    img_data = _all_images.get(sheet_name, {}).get(r_offset)
                    if img_data:
                        img_bytes, img_ext = img_data
                        valid_subs = {
                            m.get("Sub-category", "").strip()
                            for m in cat_master if m.get("Sub-category")
                        } - {""}
                        v_sub, v_conf, v_why = _img_mod.classify_image(
                            img_bytes,
                            el.master_category,
                            valid_subs,
                            description=el.raw_description,
                            image_ext=img_ext,
                            api_key=vision_api_key,
                        )
                        _vision_calls += 1
                        el.vision_subcategory = v_sub
                        el.vision_confidence  = v_conf
                        el.vision_reason      = v_why
                        # Accept vision result when:
                        #   (a) Engine 1 found no sub-category, or
                        #   (b) Vision is ≥10 points more confident
                        if v_sub and (
                            not el.master_subcategory
                            or v_conf >= clf.confidence + 10
                        ):
                            el.master_subcategory        = v_sub
                            el.classification_confidence = v_conf
                            el.classification_method     = "vision"
                            el.classification_signals   += f" | vision({v_why})"

                # ── ENGINE 2: match against BOQ (skipped when run_match=False) ─
                if run_match and budget:
                    # Engine 1 may have overridden the category (e.g. Wall Tiles → CIVIL).
                    # Re-fetch the BOQ bucket for the (possibly new) category.
                    e1_cat = (clf.category or el.detected_category or "").upper()
                    if e1_cat != cat_upper:
                        e2_budget = budget_by_cat.get(e1_cat, [])
                        e2_exists = len(e2_budget) > 0
                    else:
                        e2_budget, e2_exists = cat_budget, cat_exists

                    gfc_row_dict = {
                        "category":    clf.category or el.detected_category,
                        "subcategory": clf.subcategory,   # Engine 1 result — primary signal
                        "sub_hint":    el.detected_sub_hint,
                        "description": el.raw_description,
                        "areas":       el.raw_areas,
                        "product_name":cell("product_name"),
                        "brand":       el.raw_brand,
                        "uom":         el.norm_uom,
                        "size":        el.raw_size,
                        "finish_code": el.raw_finish_code,
                        "master_item": el.master_item,
                    }
                    matched, score, status, breakdown = best_match(
                        gfc_row_dict, e2_budget, e2_exists)

                    el.match_status    = status
                    el.match_score     = score
                    el.score_breakdown = "; ".join(f"{k}: {v}" for k, v in breakdown.items())
                    if matched:
                        el.matched_boq_no       = matched["BOQ No"]
                        el.matched_description  = matched["Description"]
                        el.matched_brand        = matched["Make/Brand"]
                        el.matched_uom          = matched["UoM"]
                        el.matched_bcs_rate     = matched["BCS Rate"]
                else:
                    el.match_status = "— (Engine 1 only)"

                # PR readiness (GF5.1).  Uses el.match_status so it is correct
                # whether or not the Engine 2 matching pass ran.
                blockers = []
                if el.match_status != "🟢 Matched":
                    blockers.append(f"match={el.match_status}")
                if el.norm_status != "Client Approved":
                    blockers.append(f"status={el.norm_status}")
                if not (el.raw_qty_ops or el.raw_qty_design):
                    blockers.append("no qty")
                if not el.norm_uom:
                    blockers.append("no uom")
                el.pr_ready   = len(blockers) == 0
                el.pr_blocker = "" if el.pr_ready else "; ".join(blockers)

                results.append(el)

    return results
